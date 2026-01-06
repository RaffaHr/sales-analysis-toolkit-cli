from __future__ import annotations

from pathlib import Path
from typing import Callable, Iterable, Optional

import hashlib
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook

COLUMN_MAP = {
    "ANO_MES": "ano_mes",
    "NR_NOTA_FISCAL": "nr_nota_fiscal",
    "CATEGORIA": "categoria",
    "CD_PRODUTO": "cd_produto",
    "DS_PRODUTO": "ds_produto",
    "Qtd de pedido": "qtd_pedidos",
    "Qtd de sku no pedido": "qtd_sku",
    "ROB": "rob",
    "Preco vendido": "preco_vendido",
    "Perc Margem Bruta% RBLD": "perc_margem_bruta",
    "Custo do produto": "custo_produto",
    "Qtd Produto Devolvido": "qtd_devolvido",
    "Devolução Receita Bruta Tot$": "devolucao_receita_bruta",
}

NUMERIC_COLUMNS = [
    "qtd_pedidos",
    "qtd_sku",
    "rob",
    "preco_vendido",
    "perc_margem_bruta",
    "custo_produto",
    "qtd_devolvido",
    "devolucao_receita_bruta",
]

PERCENT_COLUMNS = ["perc_margem_bruta"]


ProgressCallback = Callable[[int, int], None]


class SalesDataLoader:
    """Centraliza a leitura, combinação e tratamento inicial das abas de vendas."""

    def __init__(
        self,
        excel_path: Path | str,
        sheet_name: str = "VENDA",
        cache_dir: Path | str | None = Path(".cache"),
        enable_cache: bool = True,
        progress_callback: ProgressCallback | None = None,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.cache_dir = Path(cache_dir) if cache_dir is not None else None
        self.enable_cache = enable_cache
        self.progress_callback = progress_callback

    def load(self) -> pd.DataFrame:
        """Carrega as abas de vendas e devolve um DataFrame padronizado."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.excel_path}")

        sheet_names = self._resolve_sheet_names()
        cache = self._try_load_cache(sheet_names)
        if cache is not None:
            self._notify_progress(1, 1)
            return cache

        if self.progress_callback is None:
            df = self._read_with_pandas(sheet_names)
        else:
            df = self._read_with_progress(sheet_names)
        df = self._coerce_numeric(df)
        df = self._enrich(df)
        self._store_cache(df, sheet_names)
        return df

    def _resolve_sheet_names(self) -> list[str]:
        with pd.ExcelFile(self.excel_path, engine="openpyxl") as workbook:
            available = workbook.sheet_names

        if self.sheet_name in available:
            return [self.sheet_name]

        matches = [name for name in available if name.startswith(self.sheet_name)]
        if matches:
            return sorted(matches, key=_natural_sort_key)

        raise ValueError(
            "Nenhuma aba corresponde ao padrão solicitado. "
            f"Informe uma aba existente ou use prefixos como '{self.sheet_name}01'."
        )

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        available_map = {original: COLUMN_MAP[original] for original in df.columns if original in COLUMN_MAP}
        df = df.rename(columns=available_map)
        df.columns = [col.strip().lower() for col in df.columns]
        return df

    def _notify_progress(self, processed: int, total: int) -> None:
        if self.progress_callback is None:
            return
        try:
            self.progress_callback(processed, total)
        except Exception:
            # Evita que uma falha na atualização de progresso interrompa o carregamento
            pass

    def _try_load_cache(self, sheet_names: Iterable[str]) -> Optional[pd.DataFrame]:
        if not self.enable_cache or self.cache_dir is None:
            return None
        cache_path = self._cache_path(sheet_names)
        if not cache_path.exists():
            return None
        cache_mtime = cache_path.stat().st_mtime
        excel_mtime = self.excel_path.stat().st_mtime
        if cache_mtime < excel_mtime:
            return None
        try:
            return pd.read_pickle(cache_path)
        except Exception:
            return None

    def _store_cache(self, df: pd.DataFrame, sheet_names: Iterable[str]) -> None:
        if not self.enable_cache or self.cache_dir is None:
            return
        cache_path = self._cache_path(sheet_names)
        try:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_pickle(cache_path)
        except Exception:
            pass

    def _cache_path(self, sheet_names: Iterable[str]) -> Path:
        signature = ",".join(sorted(sheet_names))
        digest = hashlib.md5(signature.encode("utf-8"), usedforsecurity=False).hexdigest()[:10]
        name = f"{self.excel_path.stem}_{digest}.pkl"
        base_dir = self.cache_dir if self.cache_dir is not None else self.excel_path.parent
        return base_dir / name

    def _read_with_pandas(self, sheet_names: list[str]) -> pd.DataFrame:
        frames = []
        for name in sheet_names:
            raw = pd.read_excel(
                self.excel_path,
                sheet_name=name,
                engine="openpyxl",
                dtype={
                    "CD_PRODUTO": str,
                    "DS_PRODUTO": str,
                    "ANO_MES": str,
                    "NR_NOTA_FISCAL": str,
                },
            )
            frames.append(self._normalize_columns(raw))
        return pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]

    def _read_with_progress(self, sheet_names: list[str]) -> pd.DataFrame:
        workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        try:
            sheet_totals = {name: self._sheet_data_rows(workbook[name]) for name in sheet_names}
            total_rows = sum(sheet_totals.values())
            total_rows = max(total_rows, 1)
            self._notify_progress(0, total_rows)

            frames: list[pd.DataFrame] = []
            processed = 0
            for name in sheet_names:
                ws = workbook[name]
                header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
                try:
                    header_tuple = next(header_iter)
                except StopIteration:
                    continue
                header = [self._ensure_str(cell) for cell in header_tuple]
                data_rows: list[list[object]] = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row is None:
                        continue
                    row_values = list(row)
                    if not any(value is not None for value in row_values):
                        continue
                    if len(row_values) < len(header):
                        row_values.extend([None] * (len(header) - len(row_values)))
                    elif len(row_values) > len(header):
                        row_values = row_values[: len(header)]
                    data_rows.append(row_values)
                    processed += 1
                    self._notify_progress(processed, total_rows)
                if data_rows:
                    frame = pd.DataFrame.from_records(data_rows, columns=header)
                else:
                    frame = pd.DataFrame(columns=header)
                frames.append(self._normalize_columns(frame))

            if not frames:
                self._notify_progress(total_rows, total_rows)
                return pd.DataFrame()

            df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
            self._notify_progress(total_rows, total_rows)
            return df
        finally:
            workbook.close()

    @staticmethod
    def _sheet_data_rows(ws) -> int:
        max_row = ws.max_row or 0
        return max(max_row - 1, 0)

    @staticmethod
    def _ensure_str(value: object) -> str:
        if value is None:
            return ""
        return str(value)

    def _coerce_numeric(self, df: pd.DataFrame) -> pd.DataFrame:
        for column in NUMERIC_COLUMNS:
            if column not in df.columns:
                continue
            series = df[column]
            if series.dtype == object:
                series = (
                    series.astype(str)
                    .str.replace("%", "", regex=False)
                    .str.replace(",", ".", regex=False)
                )
            df[column] = pd.to_numeric(series, errors="coerce")
        for column in PERCENT_COLUMNS:
            if column not in df.columns:
                continue
            mask = df[column].notna()
            df.loc[mask, column] = np.where(
                df.loc[mask, column] > 1,
                df.loc[mask, column] / 100,
                df.loc[mask, column],
            )
        return df

    def _enrich(self, df: pd.DataFrame) -> pd.DataFrame:
        if "ano_mes" in df.columns:
            cleaned = df["ano_mes"].astype(str).str.replace(" ", "", regex=False)
            cleaned = cleaned.str.replace(r"[^0-9]", "", regex=True)
            cleaned = cleaned.where(cleaned.str.len() == 6)
            df["ano_mes"] = cleaned
            df["periodo"] = pd.to_datetime(df["ano_mes"], format="%Y%m", errors="coerce").dt.to_period("M")
        else:
            df["periodo"] = pd.NaT

        if "categoria" not in df.columns:
            df["categoria"] = "Sem Categoria"
        df["categoria"] = df["categoria"].fillna("Sem Categoria").astype(str).str.strip()

        if "cd_produto" not in df.columns:
            df["cd_produto"] = ""
        df["cd_produto"] = df["cd_produto"].fillna("").astype(str).str.strip()

        if "ds_produto" not in df.columns:
            df["ds_produto"] = ""
        df["ds_produto"] = df["ds_produto"].fillna("").astype(str).str.strip()

        if "nr_nota_fiscal" not in df.columns:
            df["nr_nota_fiscal"] = ""
        df["nr_nota_fiscal"] = df["nr_nota_fiscal"].fillna("").astype(str).str.strip()

        qtd_sku = df["qtd_sku"] if "qtd_sku" in df.columns else pd.Series(0, index=df.index, dtype=float)
        preco_vendido = df["preco_vendido"] if "preco_vendido" in df.columns else pd.Series(0, index=df.index, dtype=float)
        custo_unitario = df["custo_produto"] if "custo_produto" in df.columns else pd.Series(0, index=df.index, dtype=float)
        devolvido = df["qtd_devolvido"] if "qtd_devolvido" in df.columns else pd.Series(0, index=df.index, dtype=float)
        margem = df["perc_margem_bruta"] if "perc_margem_bruta" in df.columns else pd.Series(0, index=df.index, dtype=float)
        receita = df["rob"] if "rob" in df.columns else pd.Series(0, index=df.index, dtype=float)

        qtd_sku = qtd_sku.fillna(0)
        preco_vendido = preco_vendido.fillna(0)
        custo_unitario = custo_unitario.fillna(0)
        devolvido = devolvido.fillna(0)
        margem = margem.fillna(0)
        receita = receita.fillna(0)

        df["receita_bruta_calc"] = preco_vendido * qtd_sku
        df["rob"] = receita.where(receita > 0, df["receita_bruta_calc"])
        df["custo_total"] = custo_unitario * qtd_sku
        df["lucro_bruto_estimado"] = df["receita_bruta_calc"] * margem

        base = qtd_sku.to_numpy(dtype=float)
        devol = devolvido.to_numpy(dtype=float)
        taxas = np.divide(
            devol,
            base,
            out=np.zeros_like(devol, dtype=float),
            where=base > 0,
        )
        df["taxa_devolucao"] = np.nan_to_num(taxas, nan=0.0)
        return df


def load_sales_dataset(path: Path | str, sheet_name: str = "VENDA", **kwargs) -> pd.DataFrame:
    """Atalho simples para carregar o conjunto de vendas padronizado."""
    loader = SalesDataLoader(path, sheet_name, **kwargs)
    return loader.load()


def _natural_sort_key(value: str) -> list[object]:
    parts = re.split(r"(\d+)", value)
    return [int(part) if part.isdigit() else part.lower() for part in parts]
